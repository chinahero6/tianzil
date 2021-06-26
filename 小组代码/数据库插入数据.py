# -*- coding:UTF-8 -*-

import psycopg2
import openpyxl
# Connect to an existing database


conn = psycopg2.connect(host='localhost', 
	database="examdb", user="examdbo", password="pass")

cur = conn.cursor()
cur.execute('''
DROP TABLE IF EXISTS course1;
CREATE TABLE IF NOT EXISTS course1  (
    sn       VARCHAR(20),     --序号
    no       VARCHAR(20), --课程号
    name     TEXT,        --课程名称
    teacher  TEXT,        --教师名
    data     TEXT,        --课程时间
    Staff_no VARCHAR(20),  --教工号
    time     TEXT,        --上课时间
    place    TEXT,        --上课地点
    PRIMARY KEY(sn));
''')

cur.execute('''
DROP TABLE IF EXISTS student1;
CREATE TABLE IF NOT EXISTS student1 (
    sn       VARCHAR(20),     --序号
    no       VARCHAR(20), --学号
    name     TEXT,        --姓名
    clss    TEXT,        --班级
    PRIMARY KEY(sn));
''')


cur.execute('''
DROP TABLE IF EXISTS course_grade1;
CREATE TABLE IF NOT EXISTS course_grade1 (
    stu_sn VARCHAR(20),     -- 学生序号
    cou_sn VARCHAR(20),     -- 课程序号
    grade  NUMERIC(5,2), -- 最终成绩
    PRIMARY KEY(stu_sn, cou_sn));
''')


wb=openpyxl.load_workbook(r'./数据2.xlsx')
ws=wb.active

colC=ws["C"]
colD=ws["D"]
colF=ws["F"]
colG=ws["G"]
colH=ws["H"]
colI=ws["I"]
colO=ws["O"]
colP=ws["P"]
colQ=ws["Q"]
colJ=ws["J"]
colK=ws["K"]
colL=ws["L"]






for i in range(1,133):
    sn = i
    no=1910610000+i
    name = '%s' % colC[i].value
    clss = '%s' % colD[i].value
    cur.execute('''
		INSERT INTO student1(sn, no,name,clss) VALUES (%(no)s, %(sn)s,%(name)s,%(clss)s) 
	''', {'sn':sn, 'no':no,'name':name,"clss":clss} )

for i in range(1,31):
    sn = i
    no= '%s' % colF[i].value
    name = '%s' % colG[i].value
    teacher = '%s' % colH[i].value
    data = '%s' % colI[i].value
    Staff_no='%s' % colJ[i].value
    time='%s' % colK[i].value
    place='%s' % colL[i].value
    cur.execute('''
		INSERT INTO course1(sn, no,name,teacher,data,Staff_no,time,place) VALUES (%(sn)s, %(no)s,%(name)s,%(teacher)s,%(data)s,%(Staff_no)s,%(time)s,%(place)s) 
	''', {'sn':sn, 'no':no,'name':name,"teacher":teacher,"data":data,"Staff_no":Staff_no,"time":time,"place":place} )

for i in range(1,len(colO)):
    stu_sn = '%s' % colO[i].value
    cou_sn= '%s' % colP[i].value
    grade='%s' % colQ[i].value
    cur.execute('''
		INSERT INTO course_grade1(stu_sn,cou_sn,grade) VALUES (%(stu_sn)s,%(cou_sn)s,%(grade)s) 
	''', {'stu_sn':stu_sn, 'cou_sn':cou_sn,'grade':grade} )



conn.commit()